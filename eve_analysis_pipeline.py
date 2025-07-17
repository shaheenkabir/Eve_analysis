###-----------Onal_lab-------------###
##----------Bilkent University----------##
#---------Ankara, Turkey----------#

import pandas as pd
import numpy as np
import os
import re
import glob
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter, find_peaks
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
from scipy.stats import ttest_ind
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

# Constructing folder and file capture (Automation)
parent_folder = "/Users/shaheenkabir/ISO2"   # Change the folder name accordingly.
FOLDER_PATTERN = "ISO*"                                   # Change the pattern accordingly.
log_path = os.path.join(parent_folder, "pipeline_log.txt")
with open(log_path, "w") as log:
    log.write("=== Analysis Log -- Failures ===\n")
matched_folders = sorted([
    os.path.join(parent_folder, d)
    for d in os.listdir(parent_folder)
    if os.path.isdir(os.path.join(parent_folder, d)) and re.fullmatch(FOLDER_PATTERN.replace("*", ".*"), d)
])
# Starting parent-level analysis.
for input_folder in matched_folders:
    print(f"\n📂 Starting analysis in: {input_folder}")
    output_folder = os.path.join(input_folder, "results")
    os.makedirs(output_folder, exist_ok=True)

    summary_rows = []

    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            input_path = os.path.join(input_folder, filename)
            output_path = os.path.join(output_folder, f"processed_{filename}")
            temp_plot_path = "temp_plot.png"
            try:
                data = pd.read_excel(input_path)
                length = data.iloc[:, 0]
                intensity = data.iloc[:, 1]

                # Normalize and invert
                norm_intensity = intensity / intensity.max()
                inverted_intensity = 1 - norm_intensity
                norm_length = length / length.max()
                percent_length = norm_length * 100

                # Smooth and calculate derivatives
                smoothed = savgol_filter(inverted_intensity, 11, 2)
                diff = np.diff(smoothed, prepend=smoothed[0])
                change_points = np.diff(np.sign(diff), prepend=0)
                change_point_flags = (change_points < 0).astype(int)

                # Detect peaks
                peaks, _ = find_peaks(smoothed, distance=10, prominence=0.03)  # Optimal is 0.03 prominence.
                peak_percent_lengths = percent_length.iloc[peaks].values
                peak_values = smoothed[peaks]
                valid_peak_mask = (peak_percent_lengths > 28) & (peak_percent_lengths <85) # based on mutation.
                peak_percent_lengths = peak_percent_lengths[valid_peak_mask]
                peak_values = peak_values[valid_peak_mask]

                data['Normalized Intensity'] = norm_intensity
                data['Inverted Intensity'] = inverted_intensity
                data['Normalized Length'] = norm_length
                data['Percent Length'] = percent_length
                data['Smoothed'] = smoothed
                data['Difference'] = diff
                data['Change Point'] = change_point_flags
                data.to_excel(output_path, index=False)

                plt.figure(figsize=(5, 3))
                plt.plot(percent_length, smoothed, label="Smoothed", color='blue')
                plt.scatter(peak_percent_lengths, peak_values, color='red', s=50, label="Peaks > 25%")
                plt.title("Smoothed Inverted Intensity")
                plt.xlabel("Percent Length (%)")
                plt.ylabel("Inverted Intensity")
                plt.legend()
                plt.tight_layout()
                plt.savefig(temp_plot_path, dpi=200)
                plt.close()

                wb = load_workbook(output_path)
                ws = wb.active
                img = XLImage(temp_plot_path)
                img.width = 360
                img.height = 220
                ws.add_image(img, "K2")

                # Add Peaks sheet
                if "Peaks" in wb.sheetnames:
                    del wb["Peaks"]
                peak_ws = wb.create_sheet("Peaks")
                peak_ws.append(["Percent Length", "Inverted Intensity"])
                for x, y in zip(peak_percent_lengths, peak_values):
                    peak_ws.append([x, y])

                # Match peaks to change points
                change_df = data[(data["Change Point"] == 1) & (data["Percent Length"] > 28) & (data["Percent Length"] < 85)]
                closest_matches = []
                for px in peak_percent_lengths:
                    if not change_df.empty:
                        closest = change_df.iloc[(change_df["Percent Length"] - px).abs().argsort()[:1]]
                        closest_val = closest["Percent Length"].values[0]
                        closest_matches.append(closest_val)
                # Using a cutoff to discard mismatched files.
                if len(closest_matches) != 7:
                    with open(log_path, "a") as log:
                        log.write(f"❌ {filename}: Found {len(closest_matches)} stripes, not 7 → skipped.\n")
                    print(f"⚠️ Skipping {filename}: found {len(closest_matches)} stripes, not exactly 7.")
                    continue  # Skip rest of the loop

                # First 3 stripes
                stripes = closest_matches[:3]
                wb.save(output_path)
                os.remove(temp_plot_path)

                # Add Macro sheet
                macro_df = pd.DataFrame({
                    "Percent Length": percent_length,
                    "Inverted Intensity": inverted_intensity
                })
                with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    macro_df.to_excel(writer, sheet_name='Macro', index=False)

                # Add Macro_600points
                try:
                    indices = np.linspace(0, len(macro_df) - 1, 600, dtype=int)
                    squeezed_macro_df = macro_df.iloc[indices].reset_index(drop=True)
                    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        squeezed_macro_df.to_excel(writer, sheet_name='Macro_600points', index=False)
                except Exception as e:
                    print(f"❌ Error creating Macro_600points sheet for {filename}: {e}")

                summary_rows.append({
                    "File": filename,
                    "Stripe-1": stripes[0],
                    "Stripe-2": stripes[1],
                    "Stripe-3": stripes[2]
                })
            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")

    # Sort summary_rows by numeric suffix in filenames
    def extract_numeric_suffix(filename):
        nums = re.findall(r'(\d+)', filename)
        return tuple(map(int, nums)) if nums else (0,)
    summary_rows_sorted = sorted(summary_rows, key=lambda x: extract_numeric_suffix(x["File"]))
    summary_df = pd.DataFrame(summary_rows_sorted)
    summary_output_path = os.path.join(output_folder, "stripe_summary.xlsx")
    summary_df.to_excel(summary_output_path, index=False)

    all_matched_rows = []
    for row in summary_rows_sorted:
        file = row["File"]
        result_path = os.path.join(output_folder, f"processed_{file}")
        try:
            wb = load_workbook(result_path, data_only=True)
            peaks_ws = wb["Peaks"]
            peaks = []
            for i, r in enumerate(peaks_ws.iter_rows(min_row=2, values_only=True), start=2):
                if r[0] is not None:
                    peaks.append(r[0])
            df = pd.read_excel(result_path)
            change_df = df[(df["Change Point"] == 1) & (df["Percent Length"] > 28)]

            matched_stripes = []
            for px in peaks:
                if not change_df.empty:
                    closest = change_df.iloc[(change_df["Percent Length"] - px).abs().argsort()[:1]]
                    matched_val = closest["Percent Length"].values[0]
                    matched_stripes.append(matched_val)
                else:
                    matched_stripes.append(None)

            all_matched_rows.append({
                "File": file,
                "Stripe-1": matched_stripes[0],
                "Stripe-2": matched_stripes[1],
                "Stripe-3": matched_stripes[2],
                "Stripe-4": matched_stripes[3],
                "Stripe-5": matched_stripes[4],
                "Stripe-6": matched_stripes[5],
                "Stripe-7": matched_stripes[6]
            })
        except Exception as e:
            print(f"⚠️ Skipping peak matching for {file}: {e}")
    matched_all_df = pd.DataFrame(all_matched_rows)
    with pd.ExcelWriter(summary_output_path, engine='openpyxl', mode='a') as writer:
        matched_all_df.to_excel(writer, sheet_name="7_Stripes", index=False)

    wb = load_workbook(summary_output_path)
    df_7 = pd.read_excel(summary_output_path, sheet_name="7_Stripes")
    stripe_groups = {
        "Stripes_1-2": ["Stripe-1", "Stripe-2"],
        "Stripes_3-4": ["Stripe-3", "Stripe-4"],
        "Stripes_5-7": ["Stripe-5", "Stripe-6", "Stripe-7"]
    }
    with pd.ExcelWriter(summary_output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, cols in stripe_groups.items():
            group_df = df_7[["File"] + cols].copy()
            avg_row = pd.DataFrame([["AVERAGE"] + group_df[cols].mean().tolist()],
                                   columns=group_df.columns)
            stdev_row = pd.DataFrame([["STDEV.S"] + group_df[cols].std(ddof=1).tolist()],
                                     columns=group_df.columns)
            blank_row = pd.DataFrame([[None] * len(group_df.columns)], columns=group_df.columns)
            final_df = pd.concat([group_df, blank_row, avg_row, stdev_row], ignore_index=True)
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(summary_output_path)
    ws = wb.active
    cols = list(range(2, 5))  # Columns B, C, D
    data = []
    n = ws.max_row
    for col in cols:
        col_data = []
        for row in range(2, n + 1):
            val = ws.cell(row=row, column=col).value
            try:
                val = float(val)
                col_data.append(val)
            except (TypeError, ValueError):
                pass
        data.append(col_data)
    averages = [round(np.mean(col_data), 4) if len(col_data) > 0 else None for col_data in data]
    stdevs = [round(np.std(col_data, ddof=1), 4) if len(col_data) > 1 else None for col_data in data]
    ws.insert_rows(n + 1)
    ws.cell(row=n + 2, column=1, value="AVERAGE")
    ws.cell(row=n + 3, column=1, value="STDEV.S")
    for idx, col in enumerate(cols):
        ws.cell(row=n + 2, column=col, value=averages[idx])
        ws.cell(row=n + 3, column=col, value=stdevs[idx])
    wb.save(summary_output_path)
    print(f"\n📊 Stripe summary written to: {summary_output_path}")

    # Combine all Macro_600points into one Excel
    macro_percent_cols = []
    macro_intensity_cols = []
    sample_names = []
    processed_files = sorted(
        [f for f in os.listdir(output_folder) if f.startswith("processed_") and f.endswith(".xlsx")],
        key=lambda x: int(re.findall(r'(\d+)', x)[-1])
    )
    for idx, fname in enumerate(processed_files, start=1):
        fpath = os.path.join(output_folder, fname)
        try:
            df = pd.read_excel(fpath, sheet_name="Macro")
            percent_col = df.iloc[:, 0] / df.iloc[:, 0].max()
            intensity_col = df.iloc[:, 1]
            indices = np.linspace(0, len(df) - 1, 600, dtype=int)
            percent_squeezed = percent_col.iloc[indices].reset_index(drop=True)
            intensity_squeezed = intensity_col.iloc[indices].reset_index(drop=True)
            label = os.path.basename(input_folder) + f"_{idx}"
            sample_names.append(label)
            macro_percent_cols.append(percent_squeezed.rename(label))
            macro_intensity_cols.append(intensity_squeezed.rename(label))
        except Exception as e:
            print(f"❌ Skipping {fname}: {e}")

    percent_df = pd.concat(macro_percent_cols, axis=1)
    intensity_df = pd.concat(macro_intensity_cols, axis=1)
    percent_df[""] = ""
    intensity_df[""] = ""
    percent_df["Average"] = percent_df.select_dtypes(include=[np.number]).mean(axis=1)
    intensity_df["Average"] = intensity_df.select_dtypes(include=[np.number]).mean(axis=1)
    combined_macro_path = os.path.join(output_folder, "combined_macro_600points.xlsx")
    with pd.ExcelWriter(combined_macro_path) as writer:
        percent_df.to_excel(writer, sheet_name="Percent Length", index=False)
        intensity_df.to_excel(writer, sheet_name="Inverted Intensity", index=False)

# Creating all average summary file.
output_combined_file = os.path.join(parent_folder, "all_avg_summary.xlsx")
length_averages = []
intensity_averages = []
labels = []
for subdir in matched_folders:
    combined_path = os.path.join(subdir, "results", "combined_macro_600points.xlsx")
    if os.path.exists(combined_path):
        try:
            percent_df = pd.read_excel(combined_path, sheet_name="Percent Length")
            intensity_df = pd.read_excel(combined_path, sheet_name="Inverted Intensity")
            avg_length = percent_df["Average"]
            avg_intensity = intensity_df["Average"]

            name = os.path.basename(subdir)
            labels.append(name)
            length_averages.append(avg_length.rename(name))
            intensity_averages.append(avg_intensity.rename(name))
        except Exception as e:
            print(f"⚠️ Could not extract from {subdir}: {e}")
    else:
        print(f"❌ Missing file in: {subdir}")

length_avg_df = pd.concat(length_averages, axis=1)
intensity_avg_df = pd.concat(intensity_averages, axis=1)
output_combined_file = os.path.join(parent_folder, "all_avg_summary.xlsx")
if os.path.exists(output_combined_file):
    mode = "a"
    if_sheet_exists = "replace"
else:
    mode = "w"
    if_sheet_exists = None
with pd.ExcelWriter(output_combined_file, engine="openpyxl", mode=mode, if_sheet_exists= if_sheet_exists) as writer:
    length_avg_df.to_excel(writer, sheet_name="Average Percent Length", index=False)
    intensity_avg_df.to_excel(writer, sheet_name="Average Intensity", index=False)

# Plot all average curves
plt.figure(figsize=(7, 5))
for label, x, y in zip(labels, length_averages, intensity_averages):
    plt.plot(x, y, label=label)
plt.xlabel("Percent Length (%)")
plt.ylabel("Inverted Intensity")
plt.title("Average Intensity Curves")
plt.legend()
plt.tight_layout()
plot_path = os.path.join(parent_folder, "average_plot.png")
plt.savefig(plot_path, dpi=300)
plt.close()

# Folder map
folders = glob.glob(os.path.join(parent_folder, FOLDER_PATTERN))
folder_map = {}
for folder in folders:
    folder_name = os.path.basename(folder)
    key = folder_name.split("_")[-1]
    stripe_file = os.path.join(folder, "results", "stripe_summary.xlsx")
    folder_map[key] = stripe_file
print("✅ folder_map:", folder_map)

def load_stripe_column(folder_key, stripe_n):
    path = folder_map[folder_key]
    if not os.path.exists(path):
        print(f"⚠️ File missing for folder {folder_key}: {path}")
        return np.array([])
    try:
        df = pd.read_excel(path, sheet_name="7_Stripes")
    except Exception as e:
        print(f"⚠️ Could not load '7_Stripes' sheet from {path}: {e}")
        return np.array([])

    col_name = f"Stripe-{stripe_n}"
    if col_name in df.columns:
        return df[col_name].dropna().values
    else:
        print(f"⚠️ Column '{col_name}' not found in '7_Stripes' sheet of {path}")
        return np.array([])

required_groups = ["22", "25", "29"]
missing = []

for grp in required_groups:
    pattern = os.path.join(parent_folder, f"*_{grp}")
    matches = glob.glob(pattern)
    if not matches:
        missing.append(f"No match for pattern: {pattern}")
if missing:
    print("❌ One or more required groups are missing!")
    for m in missing:
        print(f"   ➜ Missing: {m}")
    print("⛔️ Aborting comparisons.")
    exit(1)  # or sys.exit(1)

# Comparison
comparisons = [("22", "25"), ("25", "29"), ("22", "29")]
for group1, group2 in comparisons:
    output_path = os.path.join(parent_folder, f"{group1}_vs_{group2}.xlsx")
    sheets_added = 0
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        for stripe_n in range(1, 8):
            data1 = load_stripe_column(group1, stripe_n)
            data2 = load_stripe_column(group2, stripe_n)

            if len(data1) < 2 or len(data2) < 2:
                print(f"⚠️ Not enough data for Stripe-{stripe_n} comparison {group1} vs {group2}, skipping")
                continue
            df = pd.DataFrame({
                f"{group1}": pd.Series(data1),
                f"{group2}": pd.Series(data2)
            })
            blank_row = pd.DataFrame([[None, None]], columns=df.columns)
            t_stat, p_val = ttest_ind(data1, data2, equal_var=False, alternative='two-sided')

            t_stat_row = pd.DataFrame([["t-stat", t_stat]], columns=df.columns)
            p_val_row = pd.DataFrame([["p-value", p_val]], columns=df.columns)
            df = pd.concat([df, blank_row, t_stat_row, p_val_row], ignore_index=True)
            df.to_excel(writer, sheet_name=f"Stripe-{stripe_n}", index=False)

            workbook = writer.book
            worksheet = writer.sheets[f"Stripe-{stripe_n}"]
            sig_fmt = workbook.add_format({'font_color': 'green'})
            not_sig_fmt = workbook.add_format({'font_color': 'red'})
            sig_text = "Significant" if p_val < 0.05 else "Not Significant"
            sig_format = sig_fmt if p_val < 0.05 else not_sig_fmt
            sig_row = len(df) + 2
            worksheet.write(sig_row-1 , 1, sig_text, sig_format)
            sheets_added += 1
    if sheets_added == 0:
        print(f"❌ No sheets added for {group1}_vs_{group2}, deleting empty file.")
        try:
            os.remove(output_path)
        except FileNotFoundError:
            pass

print("🎉 All analyses done!")
print("Good Luck!")
